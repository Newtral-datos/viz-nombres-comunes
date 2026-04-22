"""
Procesa los Excel de nacimientos (2002-2023) y genera public/data.json.
Ejecutar desde la raíz del proyecto nombres-nacimientos-viz/:
  python3 scripts/process_data.py
"""
import json, sys
from pathlib import Path
import openpyxl

RAW_DIR  = Path('../raw/nacimientos')
OUT_FILE = Path('public/data.json')

YEARS = list(range(2002, 2024))

def parse_year(year):
    suffix = str(year)[2:]
    path   = RAW_DIR / f'nomnac{suffix}.xlsx'
    wb     = openpyxl.load_workbook(path, read_only=True)
    sheet  = next(s for s in wb.sheetnames if s.upper() == 'TOTAL')
    ws     = wb[sheet]
    rows   = list(ws.iter_rows(values_only=True))
    data_h, data_m = [], []
    for r in rows:
        if r[0] == 'TOTAL':
            continue
        if r[0] and isinstance(r[0], str) and isinstance(r[1], (int, float)):
            data_h.append((r[0].strip(), int(r[1])))
        if r[3] and isinstance(r[3], str) and isinstance(r[4], (int, float)):
            data_m.append((r[3].strip(), int(r[4])))
    return data_h, data_m

# Recopilar datos
all_h = {}   # {nombre: {año: count}}
all_m = {}
totals_h = {}
totals_m = {}

for year in YEARS:
    print(f'  {year}…', end=' ', flush=True)
    h, m = parse_year(year)

    # totales de nacimientos
    # la fila TOTAL la parseamos aparte
    suffix = str(year)[2:]
    wb   = openpyxl.load_workbook(RAW_DIR / f'nomnac{suffix}.xlsx', read_only=True)
    sht  = next(s for s in wb.sheetnames if s.upper() == 'TOTAL')
    ws   = wb[sht]
    for r in ws.iter_rows(values_only=True):
        if r[0] == 'TOTAL':
            totals_h[year] = int(r[1])
            totals_m[year] = int(r[4])
            break

    for rank, (nombre, count) in enumerate(h, 1):
        all_h.setdefault(nombre, {})
        all_h[nombre][year] = {'c': count, 'r': rank}

    for rank, (nombre, count) in enumerate(m, 1):
        all_m.setdefault(nombre, {})
        all_m[nombre][year] = {'c': count, 'r': rank}

print()

# Convertir a listas indexadas por año para el frontend
def to_series(data_dict, years):
    out = {}
    for nombre, by_year in data_dict.items():
        counts = []
        ranks  = []
        for y in years:
            entry = by_year.get(y)
            counts.append(entry['c'] if entry else None)
            ranks.append(entry['r'] if entry else None)
        out[nombre] = {'counts': counts, 'ranks': ranks}
    return out

result = {
    'years':   YEARS,
    'H':       to_series(all_h, YEARS),
    'M':       to_series(all_m, YEARS),
    'totals':  {'H': totals_h, 'M': totals_m},
}

OUT_FILE.parent.mkdir(exist_ok=True)
with open(OUT_FILE, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, separators=(',', ':'))

size = OUT_FILE.stat().st_size
print(f'Guardado {OUT_FILE} ({size/1024:.0f} KB)')
print(f'Nombres únicos H: {len(result["H"])}  M: {len(result["M"])}')

# --- Generar ages.json (frecuencia y edad media del padrón 2024) ---
AGES_RAW  = Path('../raw/nombres_por_edad_media.xlsx')
AGES_FILE = Path('public/ages.json')

def _is_data_row(r):
    if r[0] is None or r[1] is None:
        return False
    try:
        int(r[0])
        return isinstance(r[1], str)
    except (ValueError, TypeError):
        return False

wb_a = openpyxl.load_workbook(AGES_RAW, read_only=True)
ages_out = {}
for sex_key, sheet_name in [('H', 'Hombres'), ('M', 'Mujeres')]:
    ages_out[sex_key] = {}
    ws = wb_a[sheet_name]
    for row in ws.iter_rows(values_only=True):
        if _is_data_row(row):
            nombre  = row[1].strip().upper()
            freq    = int(row[2]) if isinstance(row[2], (int, float)) else 0
            avg_age = round(float(row[3]), 1) if isinstance(row[3], (int, float)) else None
            ages_out[sex_key][nombre] = {'freq': freq, 'avg_age': avg_age}

with open(AGES_FILE, 'w', encoding='utf-8') as f:
    json.dump(ages_out, f, ensure_ascii=False, separators=(',', ':'))

size_a = AGES_FILE.stat().st_size
print(f'Guardado {AGES_FILE} ({size_a/1024:.0f} KB)')
print(f'Nombres únicos en padrón H: {len(ages_out["H"])}  M: {len(ages_out["M"])}')
